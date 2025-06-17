import time

import openpyxl
import pytest
from appium import webdriver
from appium.options.android import UiAutomator2Options
from appium.webdriver.appium_service import AppiumService
from appium.webdriver.common.appiumby import AppiumBy

#https://monfared.medium.com/gestures-in-appium-part2-tap-double-tap-multi-finger-6d89451260cf#7c85


def get_data():
  #  return [

   #     ["Delhi", "India"],
    #    ["Dubai","UAE"],

    #]

    workbook = openpyxl.load_workbook("..//Excel//Login.xlsx")
    sheet = workbook["login"]
    totalrows = sheet.max_row
    totalcols = sheet.max_column
    mainList =[]

    for i in range(2, totalrows+1):
        dataList = []
        for j in range(1, totalcols+1):
            data = sheet.cell(row=i,column=j).value
            dataList.insert(j, data)
        mainList.insert(i, dataList)
    return  mainList


@pytest.mark.parametrize("username","password", get_data())
def testdologin(username,password):

    desired_caps = dict(

        deviceName='ca1f79ac',
        platformName='Android',
        browserName='Chrome',
        automationName='UIAutomator2'

    )

    appium_service = AppiumService()
    appium_service.start()

    print(appium_service.is_running)
    print(appium_service.is_listening)


    # bind these caps to uiautomator2 options

    capabilities_options = UiAutomator2Options().load_capabilities(desired_caps)
    driver = webdriver.Remote('http://127.0.0.1:4723', options = capabilities_options)
    time.sleep(3)
    driver.get('https://opensource-demo.orangehrmlive.com/web/index.php/auth/login')
    time.sleep(3)
    driver.find_element(AppiumBy.XPATH, "//input[@placeholder='Username']").send_keys(username)
    time.sleep(2)
    driver.find_element(AppiumBy.XPATH, "//input[@placeholder='Password']").send_keys(password)
    time.sleep(2)
    driver.find_element(AppiumBy.XPATH, "//button[normalize-space()='Login']").click()
    appium_service.stop



